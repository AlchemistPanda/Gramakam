import sys
import subprocess
import pathlib

# ensure required packages
reqs = ["openpyxl", "requests", "beautifulsoup4", "python-slugify"]
for r in reqs:
    try:
        __import__(r.replace('-', '_'))
    except Exception:
        subprocess.check_call([sys.executable, "-m", "pip", "install", r])

from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from slugify import slugify

excel_path = r"D:\\Gramakam 2026\\docs\\Book1_Manglish.xlsx"
p = pathlib.Path(excel_path)
if not p.exists():
    print("ERROR: Excel file not found:", excel_path)
    sys.exit(1)

wb = load_workbook(excel_path)
sheet = wb.active

# Use column C (3) for Malayalam titles and write Manglish to the next empty column
INPUT_COL = 3
OUTPUT_COL = sheet.max_column + 1
MAX = sheet.max_row
sheet.cell(row=1, column=OUTPUT_COL).value = "Manglish"

session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0 (compatible)"})

def search_dcbooks(title):
    q = title.strip()
    if not q:
        return ""
    url = f"https://www.dcbooks.com/?s={requests.utils.quote(q)}"
    try:
        r = session.get(url, timeout=15)
        if r.status_code != 200:
            print(f"Warning: status {r.status_code} for query: {q}")
            return ""
        soup = BeautifulSoup(r.text, "html.parser")
        selectors = [
            "h2.entry-title a",
            "h3.entry-title a",
            "h2 a",
            ".product_title.entry-title",
            ".woocommerce-loop-product__title",
            "a[href*='/product/']"
        ]
        for sel in selectors:
            el = soup.select_one(sel)
            if el and el.get_text(strip=True):
                txt = el.get_text(strip=True)
                if any('A' <= c <= 'z' for c in txt):
                    return txt
                href = el.get('href')
                if href:
                    try:
                        r2 = session.get(href, timeout=15)
                        s2 = BeautifulSoup(r2.text, "html.parser")
                        for sel2 in ["h1.entry-title", "h1.product_title", "h1"]:
                            el2 = s2.select_one(sel2)
                            if el2 and el2.get_text(strip=True):
                                txt2 = el2.get_text(strip=True)
                                if any('A' <= c <= 'z' for c in txt2):
                                    return txt2
                    except Exception:
                        pass
        return ""
    except Exception as e:
        print(f"Error searching dcbooks for '{title}': {e}")
        return ""

results = []
for row in range(2, MAX+1):
    mal = sheet.cell(row=row, column=INPUT_COL).value
    if not mal:
        continue
    mal_text = str(mal).strip()
    print(f"Searching for row {row}: {mal_text}")
    manglish = search_dcbooks(mal_text)
    if not manglish:
        try_q = slugify(mal_text, lowercase=False)
        if try_q:
            manglish = search_dcbooks(try_q)
    sheet.cell(row=row, column=OUTPUT_COL).value = manglish
    results.append((row, mal_text, manglish))

out_path = p.with_name(p.stem + "_updated.xlsx")
wb.save(out_path)
print("Done. Updated workbook saved to:", out_path)
for r in results[:20]:
    print(r)
print("Total rows processed:", len(results))
